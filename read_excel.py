#!/usr/bin/env python3
"""
Simple script to read Excel files and create data structures
"""
import openpyxl

def read_customers():
    """Read customer data from Excel file"""
    customers = {}
    try:
        wb = openpyxl.load_workbook('telestar_customer_list.xlsx')
        sheet = wb.active
        
        for row in range(2, sheet.max_row + 1):  # Skip header
            phone = sheet.cell(row=row, column=1).value
            name = sheet.cell(row=row, column=2).value
            if phone and name:
                customers[phone] = name
                
        print(f"Loaded {len(customers)} customers")
        return customers
    except Exception as e:
        print(f"Error reading customers: {e}")
        return {}

def read_merchants():
    """Read merchant data from Excel file"""
    merchants = {}
    try:
        wb = openpyxl.load_workbook('merchant_list.xlsx')
        sheet = wb.active
        
        for row in range(2, sheet.max_row + 1):  # Skip header
            merchant_id = sheet.cell(row=row, column=1).value
            name = sheet.cell(row=row, column=2).value
            if merchant_id and name:
                merchants[str(merchant_id)] = name
                
        print(f"Loaded {len(merchants)} merchants")
        return merchants
    except Exception as e:
        print(f"Error reading merchants: {e}")
        return {}

if __name__ == "__main__":
    customers = read_customers()
    merchants = read_merchants()
    
    print("\n=== SAMPLE CUSTOMERS ===")
    for phone, name in list(customers.items())[:5]:
        print(f"{phone}: {name}")
    
    print("\n=== SAMPLE MERCHANTS ===")
    for mid, name in list(merchants.items())[:5]:
        print(f"{mid}: {name}")
